import tempfile
from django.contrib import messages
from django.contrib.auth import login, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.management import call_command
from django.db import IntegrityError, transaction
from django.db.models import ProtectedError
from django.db.models import Count, Max, Q
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from openpyxl import Workbook, load_workbook

from .decorators import app_admin_required, checker_required, teacher_required, is_app_admin, is_teacher_user
from .forms import (
    AdminCorrectionReviewForm,
    BackupUploadForm,
    AdminTeacherProgressEditForm,
    ClassSetupForm,
    CorrectionRequestForm,
    LockRecordForm,
    PasswordResetByAdminForm,
    SelectCheckForm,
    StudentForm,
    StudentImportUploadForm,
    TeacherCourseAssignmentForm,
    TeacherCompleteForm,
    TeacherProgressForm,
    IssueWeekForm,
    UserCreateByAdminForm,
    UserDeleteConfirmForm,
    UserProfileForm,
)
from .models import (
    ClassRoom,
    ClassSubject,
    ClassSubjectChapter,
    CopyCheckRecord,
    CorrectionRequest,
    DailyBackup,
    Student,
    Subject,
    TeacherCourseAssignment,
    TeacherCourseProgress,
    TeacherProgressCorrectionRequest,
    UserProfile,
)
from .utils import ensure_daily_backup, get_backup_path, log_action

PENDING_STATUS = 'pending'


def _days_since(dt):
    if not dt:
        return None
    return (timezone.localdate() - timezone.localtime(dt).date()).days


def _relative_date(dt):
    if not dt:
        return 'Never'
    days = _days_since(dt)
    if days == 0:
        return 'Today'
    if days == 1:
        return 'Yesterday'
    return f'{days} days ago'


def _progress_percent(checked, expected):
    if expected <= 0:
        return 0
    return round((checked / expected) * 100)




def _teacher_pending_progress_rows(rows):
    result = []
    for row in rows:
        row.days_pending = _days_since(row.created_at) or 0
        row.needs_attention = row.days_pending >= 3
        result.append(row)
    result.sort(key=lambda item: (-item.days_pending, item.assignment.classroom.name, item.assignment.subject.name, -item.week_no))
    return result


def _last_checker_name(record):
    if not record:
        return '—'
    return _display_user(record.entered_by)


def _class_priority(classroom):
    """Sort classes 8 to 1 first; PG/Nursery/KG and unusual names go below."""
    raw = (classroom.name or '').strip().lower().replace('grade', '').strip()
    try:
        number = int(raw)
    except (TypeError, ValueError):
        number = None
    if number is not None and 1 <= number <= 8:
        return (0, -number, classroom.section or '')
    return (1, 0, classroom.section or '', classroom.name or '')


def _student_attention_payload(student):
    last_record = CopyCheckRecord.objects.filter(student=student, locked=True).select_related(
        'class_subject__subject', 'entered_by', 'entered_by__checker_profile'
    ).order_by('-locked_at', '-id').first()
    last_checked_at = last_record.locked_at if last_record else None
    total_checks = CopyCheckRecord.objects.filter(student=student, locked=True).count()
    incomplete_count = CopyCheckRecord.objects.filter(student=student, locked=True, status=CopyCheckRecord.STATUS_INCOMPLETE).count()
    return {
        'student': student,
        'classroom': student.classroom,
        'last_record': last_record,
        'last_checked_at': last_checked_at,
        'last_checked_label': _relative_date(last_checked_at),
        'days_since': _days_since(last_checked_at),
        'last_subject': last_record.subject.name if last_record else '—',
        'last_status': last_record.get_status_display() if last_record else '—',
        'last_checker': _last_checker_name(last_record),
        'total_checks': total_checks,
        'incomplete_count': incomplete_count,
        'sort_key': (0 if last_checked_at is None else 1, last_checked_at or timezone.datetime.min.replace(tzinfo=timezone.get_current_timezone()), student.roll_no, student.full_name),
    }


def _attention_students_for_class(classroom, limit=None):
    students = Student.objects.filter(is_active=True, classroom=classroom).select_related('classroom').order_by('roll_no', 'full_name')
    rows = [_student_attention_payload(student) for student in students]
    rows.sort(key=lambda row: row['sort_key'])
    return rows[:limit] if limit else rows


def _attention_student_rows(limit=50):
    rows = []
    students = Student.objects.filter(is_active=True, classroom__is_active=True).select_related('classroom').order_by(
        'classroom__name', 'classroom__section', 'roll_no', 'full_name'
    )
    for student in students:
        row = _student_attention_payload(student)
        row['sort_key'] = (_class_priority(student.classroom), row['sort_key'])
        rows.append(row)
    rows.sort(key=lambda row: row['sort_key'])
    return rows[:limit]


def _checking_record_queryset():
    return CopyCheckRecord.objects.filter(locked=True).select_related(
        'student',
        'classroom',
        'class_subject__subject',
        'chapter',
        'entered_by',
        'entered_by__checker_profile',
        'actual_checker_user',
        'actual_checker_user__checker_profile',
    )


def _display_user(user):
    profile = getattr(user, 'checker_profile', None)
    if profile:
        return profile.display_name
    return user.username


def _subject_rows_from_post(request):
    subject_names = request.POST.getlist('subject_name')
    rows = []
    errors = []
    seen = set()
    for index, raw_name in enumerate(subject_names):
        name = (raw_name or '').strip()
        if not name:
            continue
        if name.lower() in seen:
            errors.append(f'Subject row {index + 1}: duplicate subject "{name}".')
            continue
        seen.add(name.lower())
        rows.append({'name': name, 'chapter_count': 1})
    if not rows:
        errors.append('Add at least one subject.')
    return rows, errors


def _validate_class_subject_changes(classroom, subject_rows):
    return []


def _sync_class_subjects(classroom, subject_rows):
    submitted_subject_ids = []
    for row in subject_rows:
        subject, _created = Subject.objects.get_or_create(
            name=row['name'],
            defaults={'is_active': True},
        )
        if not subject.is_active:
            subject.is_active = True
            subject.save(update_fields=['is_active'])

        class_subject, _created = ClassSubject.objects.update_or_create(
            classroom=classroom,
            subject=subject,
            defaults={'chapter_count': 1, 'is_active': True},
        )
        submitted_subject_ids.append(class_subject.id)

    # Subjects removed from the edit form are hidden, not deleted, so older records remain safe.
    ClassSubject.objects.filter(classroom=classroom).exclude(id__in=submitted_subject_ids).update(is_active=False)


def _class_subject_options(classroom):
    return ClassSubject.objects.filter(
        classroom=classroom,
        is_active=True,
        subject__is_active=True,
    ).select_related('subject').order_by('subject__name')



def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        login(request, form.get_user())
        return redirect('dashboard')
    return render(request, 'checker/login.html', {'form': form})


@login_required
def dashboard(request):
    if is_app_admin(request.user):
        today = timezone.localdate()
        month_start = today.replace(day=1)
        pending_requests = CorrectionRequest.objects.filter(status=CorrectionRequest.STATUS_PENDING).count()
        context = {
            'total_locked': CopyCheckRecord.objects.filter(locked=True).count(),
            'month_locked': CopyCheckRecord.objects.filter(locked=True, locked_at__date__gte=month_start).count(),
            'today_locked': CopyCheckRecord.objects.filter(locked=True, locked_at__date=today).count(),
            'pending_requests': pending_requests,
            'teacher_assignments_count': TeacherCourseAssignment.objects.filter(is_active=True).count(),
            'teacher_pending_count': TeacherCourseProgress.objects.filter(status=TeacherCourseProgress.STATUS_NOT_COMPLETED).count(),
            'teacher_completed_count': TeacherCourseProgress.objects.filter(status=TeacherCourseProgress.STATUS_COMPLETED).count(),
        }
        return render(request, 'checker/admin_dashboard.html', context)

    if is_teacher_user(request.user):
        assignments = TeacherCourseAssignment.objects.filter(
            teacher=request.user,
            is_active=True,
            class_subject__is_active=True,
            class_subject__classroom__is_active=True,
            class_subject__subject__is_active=True,
        ).select_related('class_subject__classroom', 'class_subject__subject').annotate(
            total_weeks=Count('progress_rows'),
            completed_weeks=Count('progress_rows', filter=Q(progress_rows__status=TeacherCourseProgress.STATUS_COMPLETED)),
            pending_weeks=Count('progress_rows', filter=Q(progress_rows__status=TeacherCourseProgress.STATUS_NOT_COMPLETED)),
            last_completed_week=Max('progress_rows__week_no', filter=Q(progress_rows__status=TeacherCourseProgress.STATUS_COMPLETED)),
        )
        assignment_rows = []
        for assignment in assignments:
            total = assignment.total_weeks or 0
            completed = assignment.completed_weeks or 0
            pending = assignment.pending_weeks or 0
            assignment_rows.append({
                'assignment': assignment,
                'total': total,
                'completed': completed,
                'not_completed': pending,
                'progress': _progress_percent(completed, total),
                'last_completed_week': assignment.last_completed_week,
            })
        assignment_rows.sort(key=lambda row: (-row['not_completed'], row['assignment'].classroom.name, row['assignment'].subject.name))
        pending_qs = TeacherCourseProgress.objects.filter(
            assignment__teacher=request.user,
            status=TeacherCourseProgress.STATUS_NOT_COMPLETED,
        )
        completed_qs = TeacherCourseProgress.objects.filter(
            assignment__teacher=request.user,
            status=TeacherCourseProgress.STATUS_COMPLETED,
        )
        context = {
            'assignment_rows': assignment_rows,
            'pending_rows': pending_qs.count(),
            'completed_rows': completed_qs.count(),
            'quick_pending_rows': _teacher_pending_progress_rows(pending_qs.select_related(
                'assignment__class_subject__classroom',
                'assignment__class_subject__subject',
            )[:6]),
        }
        return render(request, 'checker/teacher_dashboard.html', context)

    today = timezone.localdate()
    month_start = today.replace(day=1)
    context = {
        'today_locked': CopyCheckRecord.objects.filter(entered_by=request.user, locked=True, locked_at__date=today).count(),
        'month_locked': CopyCheckRecord.objects.filter(entered_by=request.user, locked=True, locked_at__date__gte=month_start).count(),
        'recent_records': _checking_record_queryset().filter(entered_by=request.user).order_by('-locked_at', '-id')[:10],
        'classes': ClassRoom.objects.filter(is_active=True).order_by('name', 'section'),
        'attention_students': _attention_student_rows(3),
    }
    return render(request, 'checker/checker_dashboard.html', context)


@login_required
def attention_list(request):
    is_admin = is_app_admin(request.user)
    show = request.GET.get('show')
    classroom_id = request.GET.get('classroom')
    selected_classroom = None
    class_rows = []

    if classroom_id:
        selected_classroom = get_object_or_404(ClassRoom, pk=classroom_id, is_active=True)
        class_limit = 10 if show == 'more' else 3
        class_rows = _attention_students_for_class(selected_classroom, class_limit)
    else:
        class_limit = 10 if show == 'more' else 3

    admin_student_rows = _attention_student_rows(50) if is_admin else []
    if is_admin and request.GET.get('export') == 'student':
        return _export_attention_students(admin_student_rows)
    if request.GET.get('export') == 'class' and selected_classroom:
        return _export_attention_class(selected_classroom, _attention_students_for_class(selected_classroom, None))
    context = {
        'limit': class_limit,
        'more_limit': 10,
        'show_more': show == 'more',
        'is_admin_view': is_admin,
        'classes': ClassRoom.objects.filter(is_active=True).order_by('name', 'section'),
        'selected_classroom': selected_classroom,
        'class_rows': class_rows,
        'student_rows': admin_student_rows,
    }
    return render(request, 'checker/attention_list.html', context)


@app_admin_required
def admin_setup_side(request):
    context = {
        'class_count': ClassRoom.objects.filter(is_active=True).count(),
        'student_count': Student.objects.filter(is_active=True).count(),
        'subject_count': Subject.objects.filter(is_active=True).count(),
        'teacher_count': UserProfile.objects.filter(role=UserProfile.ROLE_TEACHER, user__is_active=True).count(),
        'checker_count': UserProfile.objects.filter(role=UserProfile.ROLE_CHECKER, user__is_active=True).count(),
        'backup_count': DailyBackup.objects.count(),
    }
    return render(request, 'checker/admin_setup_side.html', context)


@app_admin_required
def admin_checking_side(request):
    today = timezone.localdate()
    month_start = today.replace(day=1)
    context = {
        'today_locked': CopyCheckRecord.objects.filter(locked=True, locked_at__date=today).count(),
        'month_locked': CopyCheckRecord.objects.filter(locked=True, locked_at__date__gte=month_start).count(),
        'total_locked': CopyCheckRecord.objects.filter(locked=True).count(),
        'attention_students': _attention_student_rows(3),
        'recent_records': _checking_record_queryset().order_by('-locked_at', '-id')[:10],
    }
    return render(request, 'checker/admin_checking_side.html', context)


@app_admin_required
def admin_teaching_side(request):
    context = {
        'teacher_assignments_count': TeacherCourseAssignment.objects.filter(is_active=True).count(),
        'teacher_pending_count': TeacherCourseProgress.objects.filter(status=TeacherCourseProgress.STATUS_NOT_COMPLETED).count(),
        'teacher_completed_count': TeacherCourseProgress.objects.filter(status=TeacherCourseProgress.STATUS_COMPLETED).count(),
        'next_week_display': 'Per course',
        'recent_teacher_progress': TeacherCourseProgress.objects.select_related(
            'assignment__teacher__checker_profile',
            'assignment__class_subject__classroom',
            'assignment__class_subject__subject',
        ).order_by('-updated_at')[:10],
    }
    return render(request, 'checker/admin_teaching_side.html', context)


@checker_required
def select_checking(request):
    form = SelectCheckForm(request.GET or None)
    if request.GET.get('classroom') and form.is_valid():
        classroom = form.cleaned_data['classroom']
        return redirect(f"{reverse('checking_list')}?classroom={classroom.id}")
    return render(request, 'checker/select_checking.html', {'form': form})


@checker_required
def checking_list(request):
    classroom = get_object_or_404(ClassRoom, pk=request.GET.get('classroom'), is_active=True)
    students = Student.objects.filter(classroom=classroom, is_active=True).order_by('roll_no', 'full_name')
    class_subjects = _class_subject_options(classroom)
    attention_rows = _attention_students_for_class(classroom, 10)
    recent_records = _checking_record_queryset().filter(classroom=classroom).order_by('-locked_at', '-id')[:10]
    context = {
        'classroom': classroom,
        'students': students,
        'class_subjects': class_subjects,
        'attention_rows': attention_rows,
        'recent_records': recent_records,
        'refresh_url': request.get_full_path(),
    }
    return render(request, 'checker/checking_list.html', context)


@checker_required
@require_POST
def lock_record(request):
    classroom = get_object_or_404(ClassRoom, pk=request.POST.get('classroom_id'), is_active=True)
    next_url = request.POST.get('next') or f"{reverse('checking_list')}?classroom={classroom.id}"
    save_flags = request.POST.getlist('save_row')
    student_ids = request.POST.getlist('student_id')
    class_subject_ids = request.POST.getlist('class_subject_id')
    statuses = request.POST.getlist('status')
    remarks_list = request.POST.getlist('remarks')
    actual_checker_names = request.POST.getlist('actual_checker_name')

    candidate_indexes = {int(raw) for raw in save_flags if str(raw).isdigit()}
    row_count = max(len(student_ids), len(class_subject_ids), len(statuses), len(remarks_list), len(actual_checker_names), (max(candidate_indexes) + 1 if candidate_indexes else 0))
    records_to_create = []
    errors = []

    for index in range(row_count):
        row_selected = index in candidate_indexes
        student_id = student_ids[index] if index < len(student_ids) else ''
        class_subject_id = class_subject_ids[index] if index < len(class_subject_ids) else ''
        status = statuses[index] if index < len(statuses) else CopyCheckRecord.STATUS_COMPLETE
        remarks = (remarks_list[index] if index < len(remarks_list) else '').strip()
        actual_checker_name = (actual_checker_names[index] if index < len(actual_checker_names) else '').strip()

        if not row_selected and not student_id and not class_subject_id and not remarks and not actual_checker_name:
            continue
        if not row_selected:
            continue

        row_number = index + 1
        if not student_id:
            errors.append(f'Row {row_number}: select a student.')
            continue
        if not class_subject_id:
            errors.append(f'Row {row_number}: select a subject.')
            continue
        if status not in dict(CopyCheckRecord.STATUS_CHOICES):
            errors.append(f'Row {row_number}: invalid status.')
            continue
        if status == CopyCheckRecord.STATUS_INCOMPLETE and not remarks:
            errors.append(f'Row {row_number}: details are required when status is Incomplete.')
            continue

        try:
            student = Student.objects.get(pk=student_id, classroom=classroom, is_active=True)
        except Student.DoesNotExist:
            errors.append(f'Row {row_number}: selected student does not belong to this class.')
            continue
        try:
            class_subject = ClassSubject.objects.select_related('subject').get(
                pk=class_subject_id,
                classroom=classroom,
                is_active=True,
                subject__is_active=True,
            )
        except ClassSubject.DoesNotExist:
            errors.append(f'Row {row_number}: selected subject is not assigned to this class.')
            continue

        records_to_create.append(CopyCheckRecord(
            student=student,
            classroom=classroom,
            class_subject=class_subject,
            chapter=None,
            status=status,
            remarks=remarks,
            entered_by=request.user,
            actual_checker_user=None,
            actual_checker_name=actual_checker_name,
            locked=True,
            locked_at=timezone.now(),
        ))

    if errors:
        for error in errors:
            messages.error(request, error)
        messages.error(request, 'No records were saved. Fix the row errors and try again.')
        return redirect(next_url)

    if not records_to_create:
        messages.warning(request, 'No selected rows were ready to save.')
        return redirect(next_url)

    with transaction.atomic():
        created_records = CopyCheckRecord.objects.bulk_create(records_to_create)
        for record in created_records:
            log_action(request.user, 'LOCK_RECORD', record, f'Locked flexible copy-checking record for {record.student.full_name}')

    ensure_daily_backup(reason='Checker saved copy-checking batch', user=request.user)
    messages.success(request, f'Saved and locked {len(records_to_create)} checking record(s).')
    return redirect(next_url)



@checker_required
def request_correction(request, record_id):
    messages.info(request, 'Correction requests are disabled. Add a new checking row if a copy is checked again or an old entry was mistaken.')
    return redirect('select_checking')


@app_admin_required
def class_list_admin(request):
    classes = ClassRoom.objects.filter(is_active=True).prefetch_related('class_subjects__subject', 'students')
    return render(request, 'checker/class_list_admin.html', {'classes': classes})


@app_admin_required
def class_setup(request, class_id=None):
    classroom = None
    if class_id:
        classroom = get_object_or_404(ClassRoom, pk=class_id)

    if request.method == 'POST':
        form = ClassSetupForm(request.POST, instance=classroom)
        subject_rows, subject_errors = _subject_rows_from_post(request)
        if not subject_errors:
            subject_errors.extend(_validate_class_subject_changes(classroom, subject_rows))
        if form.is_valid() and not subject_errors:
            with transaction.atomic():
                classroom = form.save()
                _sync_class_subjects(classroom, subject_rows)
                log_action(request.user, 'SAVE_CLASS_SETUP', classroom, f'Saved class setup for {classroom}')
            ensure_daily_backup(reason='Admin saved class setup', user=request.user)
            messages.success(request, 'Class and assigned subjects saved.')
            return redirect('class_list_admin')
        for error in subject_errors:
            messages.error(request, error)
    else:
        form = ClassSetupForm(instance=classroom)
        if classroom:
            subject_rows = [
                {'name': item.subject.name, 'chapter_count': 1}
                for item in classroom.class_subjects.filter(is_active=True).select_related('subject')
            ]
        else:
            subject_rows = []

    while len(subject_rows) < 8:
        subject_rows.append({'name': '', 'chapter_count': 1})

    return render(request, 'checker/class_setup.html', {
        'form': form,
        'classroom': classroom,
        'subject_rows': subject_rows,
    })


@app_admin_required
def student_create(request):
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            try:
                student = form.save()
            except IntegrityError:
                messages.error(request, 'This roll number already exists in the selected class.')
            else:
                log_action(request.user, 'ADD_STUDENT', student, f'Added student {student}')
                ensure_daily_backup(reason='Admin added a student', user=request.user)
                messages.success(request, 'Student saved.')
                return redirect('student_list_admin')
    else:
        initial = {}
        if request.GET.get('classroom'):
            initial['classroom'] = request.GET.get('classroom')
        form = StudentForm(initial=initial)
    return render(request, 'checker/student_form.html', {'form': form})


@app_admin_required
def admin_records(request):
    records = _checking_record_queryset().order_by('-locked_at', '-id')
    classroom_id = request.GET.get('classroom')
    subject_id = request.GET.get('subject')
    student_q = (request.GET.get('student') or '').strip()
    status = request.GET.get('status')
    checker_id = request.GET.get('checker')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if classroom_id:
        records = records.filter(classroom_id=classroom_id)
    if subject_id:
        records = records.filter(class_subject__subject_id=subject_id)
    if student_q:
        records = records.filter(Q(student__full_name__icontains=student_q) | Q(student__roll_no__icontains=student_q))
    if status:
        records = records.filter(status=status)
    if checker_id:
        records = records.filter(entered_by_id=checker_id)
    if date_from:
        records = records.filter(locked_at__date__gte=date_from)
    if date_to:
        records = records.filter(locked_at__date__lte=date_to)

    if request.GET.get('export') == 'xlsx':
        return _export_checking_records(records)

    return render(request, 'checker/admin_records.html', {
        'records': records[:50],
        'classes': ClassRoom.objects.filter(is_active=True).order_by('name', 'section'),
        'subjects': Subject.objects.filter(is_active=True).order_by('name'),
        'checkers': User.objects.filter(checker_profile__role=UserProfile.ROLE_CHECKER).select_related('checker_profile').order_by('checker_profile__display_name', 'username'),
        'status_choices': CopyCheckRecord.STATUS_CHOICES,
        'filters': request.GET,
    })



def _export_attention_students(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attention List'
    ws.append(['Class', 'Roll No', 'Student', 'Last Checked', 'Days Since Last Check', 'Last Subject', 'Last Status', 'Total Checks', 'Incomplete Count', 'Last Checker'])
    for row in rows:
        ws.append([
            str(row['classroom']),
            row['student'].roll_no,
            row['student'].full_name,
            row['last_checked_label'],
            row['days_since'] if row['days_since'] is not None else '',
            row['last_subject'],
            row['last_status'],
            row['total_checks'],
            row['incomplete_count'],
            row['last_checker'],
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="attention_students.xlsx"'
    wb.save(response)
    return response


def _export_attention_class(classroom, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Class Attention'
    ws.append(['Class', str(classroom)])
    ws.append([])
    ws.append(['Roll No', 'Student', 'Last Checked', 'Days Since Last Check', 'Last Subject', 'Last Status', 'Total Checks', 'Incomplete Count', 'Last Checker'])
    for row in rows:
        ws.append([
            row['student'].roll_no,
            row['student'].full_name,
            row['last_checked_label'],
            row['days_since'] if row['days_since'] is not None else '',
            row['last_subject'],
            row['last_status'],
            row['total_checks'],
            row['incomplete_count'],
            row['last_checker'],
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attention_{classroom.id}.xlsx"'
    wb.save(response)
    return response


def _export_checking_records(records):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Checking Records'
    ws.append(['Date', 'Class', 'Roll No', 'Student', 'Subject', 'Status', 'Details', 'Entered By', 'Actual Checker'])
    for r in records[:5000]:
        ws.append([
            timezone.localtime(r.locked_at).strftime('%Y-%m-%d %H:%M') if r.locked_at else '',
            str(r.classroom),
            r.student.roll_no,
            r.student.full_name,
            r.subject.name,
            r.get_status_display(),
            r.remarks,
            _display_user(r.entered_by),
            r.actual_checker_display,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="checking_records.xlsx"'
    wb.save(response)
    return response



@app_admin_required
def student_list_admin(request):
    students = Student.objects.filter(is_active=True).select_related('classroom')
    classroom_id = request.GET.get('classroom')
    q = request.GET.get('q')
    if classroom_id:
        students = students.filter(classroom_id=classroom_id)
    if q:
        students = students.filter(Q(full_name__icontains=q) | Q(roll_no__icontains=q))
    return render(request, 'checker/student_list_admin.html', {
        'students': students[:500],
        'classes': ClassRoom.objects.filter(is_active=True),
    })


def _cell_to_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_student_import_rows(uploaded_file, classroom):
    valid_rows = []
    problem_rows = []
    seen_rolls = set()

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active

    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        name = _cell_to_text(row[0] if len(row) > 0 else '')
        roll_no = _cell_to_text(row[1] if len(row) > 1 else '')

        if not name and not roll_no:
            continue

        problem = ''
        if not name:
            problem = 'Student name is missing.'
        elif not roll_no:
            problem = 'Roll no is missing.'
        elif roll_no in seen_rolls:
            problem = 'Duplicate roll no inside this Excel file.'
        elif Student.objects.filter(classroom=classroom, roll_no=roll_no).exists():
            problem = 'This roll no already exists in the selected class.'

        item = {'row_number': row_number, 'name': name, 'roll_no': roll_no}
        if problem:
            item['problem'] = problem
            problem_rows.append(item)
            continue

        seen_rolls.add(roll_no)
        valid_rows.append(item)

    return valid_rows, problem_rows


@app_admin_required
def student_import(request):
    preview = None
    result = None
    form = StudentImportUploadForm()

    if request.method == 'POST' and request.POST.get('confirm_import') == '1':
        classroom_id = request.session.get('student_import_classroom_id')
        valid_rows = request.session.get('student_import_valid_rows', [])
        classroom = get_object_or_404(ClassRoom, pk=classroom_id, is_active=True)

        imported = []
        skipped = []
        with transaction.atomic():
            for row in valid_rows:
                if Student.objects.filter(classroom=classroom, roll_no=row['roll_no']).exists():
                    skipped.append({**row, 'problem': 'This roll no already exists now, so it was not imported.'})
                    continue
                student = Student.objects.create(
                    classroom=classroom,
                    roll_no=row['roll_no'],
                    full_name=row['name'],
                )
                imported.append(student)
                log_action(request.user, 'IMPORT_STUDENT', student, f'Imported student {student}')

        request.session.pop('student_import_classroom_id', None)
        request.session.pop('student_import_valid_rows', None)
        request.session.pop('student_import_problem_rows', None)

        if imported:
            ensure_daily_backup(reason='Admin imported students', user=request.user)
            messages.success(request, f'Imported {len(imported)} student(s).')
        if skipped:
            messages.warning(request, f'{len(skipped)} row(s) were left for manual entry because of problems.')

        result = {'classroom': classroom, 'imported': imported, 'skipped': skipped}

    elif request.method == 'POST':
        form = StudentImportUploadForm(request.POST, request.FILES)
        if form.is_valid():
            classroom = form.cleaned_data['classroom']
            uploaded_file = form.cleaned_data['file']
            if not uploaded_file.name.lower().endswith('.xlsx'):
                messages.error(request, 'Please upload an .xlsx Excel file.')
            else:
                try:
                    valid_rows, problem_rows = _read_student_import_rows(uploaded_file, classroom)
                except Exception as exc:  # keep the UI clear instead of showing a traceback
                    messages.error(request, f'Could not read the Excel file: {exc}')
                else:
                    request.session['student_import_classroom_id'] = classroom.id
                    request.session['student_import_valid_rows'] = valid_rows
                    request.session['student_import_problem_rows'] = problem_rows
                    preview = {'classroom': classroom, 'valid_rows': valid_rows, 'problem_rows': problem_rows}
                    if not valid_rows:
                        messages.warning(request, 'No valid rows are ready for import. Check the problem rows below.')

    return render(request, 'checker/student_import.html', {
        'form': form,
        'preview': preview,
        'result': result,
    })


@login_required
def student_profile(request, student_id):
    student = get_object_or_404(Student.objects.select_related('classroom'), pk=student_id)
    subject_id = request.GET.get('subject')
    status = request.GET.get('status')

    records = student.copy_records.filter(locked=True).select_related(
        'class_subject__subject', 'entered_by', 'entered_by__checker_profile', 'actual_checker_user'
    ).order_by('-locked_at', '-id')
    if subject_id:
        records = records.filter(class_subject__subject_id=subject_id)
    if status:
        records = records.filter(status=status)

    all_records = student.copy_records.filter(locked=True)
    last_record = all_records.select_related('class_subject__subject').order_by('-locked_at', '-id').first()
    summary = {
        'total_checks': all_records.count(),
        'complete': all_records.filter(status=CopyCheckRecord.STATUS_COMPLETE).count(),
        'incomplete': all_records.filter(status=CopyCheckRecord.STATUS_INCOMPLETE).count(),
        'last_checked': last_record.locked_at if last_record else None,
        'last_subject': last_record.subject.name if last_record else '—',
    }
    context = {
        'student': student,
        'records': records[:300],
        'summary': summary,
        'subjects': Subject.objects.filter(class_subjects__classroom=student.classroom, class_subjects__is_active=True).distinct().order_by('name'),
        'status_choices': CopyCheckRecord.STATUS_CHOICES,
        'filters': request.GET,
    }
    if request.GET.get('export') == 'xlsx':
        return _export_student_profile(student, records)
    return render(request, 'checker/student_profile.html', context)


def _export_student_profile(student, records):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Student Profile'
    ws.append(['Student', student.full_name])
    ws.append(['Class', str(student.classroom)])
    ws.append(['Roll No', student.roll_no])
    ws.append([])
    ws.append(['Date', 'Subject', 'Status', 'Details', 'Entered By', 'Actual Checker'])
    for r in records[:5000]:
        ws.append([
            timezone.localtime(r.locked_at).strftime('%Y-%m-%d %H:%M') if r.locked_at else '',
            r.subject.name,
            r.get_status_display(),
            r.remarks,
            _display_user(r.entered_by),
            r.actual_checker_display,
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="student_{student.id}_checking.xlsx"'
    wb.save(response)
    return response



@app_admin_required
def correction_requests_admin(request):
    requests_qs = CorrectionRequest.objects.select_related(
        'record', 'record__student', 'record__class_subject__subject', 'record__chapter', 'requested_by'
    )
    status = request.GET.get('status') or CorrectionRequest.STATUS_PENDING
    if status:
        requests_qs = requests_qs.filter(status=status)
    return render(request, 'checker/correction_requests_admin.html', {
        'correction_requests': requests_qs[:300],
        'status': status,
        'status_choices': CorrectionRequest.STATUS_CHOICES,
    })


@app_admin_required
def review_correction_request(request, request_id):
    correction = get_object_or_404(CorrectionRequest.objects.select_related('record'), pk=request_id)
    record = correction.record
    if request.method == 'POST':
        form = AdminCorrectionReviewForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            correction.admin_note = form.cleaned_data['admin_note']
            correction.reviewed_by = request.user
            correction.reviewed_at = timezone.now()

            if action == AdminCorrectionReviewForm.ACTION_UNLOCK:
                record.locked = False
                record.locked_at = None
                record.save(update_fields=['locked', 'locked_at', 'updated_at'])
                correction.status = CorrectionRequest.STATUS_APPROVED
                log_action(request.user, 'UNLOCK_RECORD', record, f'Unlocked record after correction request #{correction.pk}')
                messages.success(request, 'Record unlocked. Checker can now edit and lock it again.')

            elif action == AdminCorrectionReviewForm.ACTION_CORRECT:
                if form.cleaned_data['status']:
                    record.status = form.cleaned_data['status']
                record.remarks = form.cleaned_data['remarks']
                record.actual_checker_user = None
                record.actual_checker_name = form.cleaned_data['actual_checker_name']
                record.save(update_fields=['status', 'remarks', 'actual_checker_user', 'actual_checker_name', 'updated_at'])
                correction.status = CorrectionRequest.STATUS_APPROVED
                log_action(request.user, 'CORRECT_RECORD', record, f'Admin corrected record after request #{correction.pk}')
                messages.success(request, 'Correction applied by admin.')

            else:
                correction.status = CorrectionRequest.STATUS_REJECTED
                log_action(request.user, 'REJECT_CORRECTION', correction, f'Rejected correction request #{correction.pk}')
                messages.info(request, 'Correction request rejected.')

            correction.save()
            ensure_daily_backup(reason='Admin reviewed a correction request', user=request.user)
            return redirect('correction_requests_admin')
    else:
        form = AdminCorrectionReviewForm(initial={
            'action': AdminCorrectionReviewForm.ACTION_CORRECT,
            'status': correction.requested_status or record.status,
            'remarks': correction.requested_remarks or record.remarks,
            'actual_checker_name': correction.requested_actual_checker_name or record.actual_checker_name,
        })
    return render(request, 'checker/review_correction_request.html', {'correction': correction, 'record': record, 'form': form})




def _force_delete_user(user):
    """Delete a user and related app-owned records. The currently logged-in admin is protected in the view."""
    CorrectionRequest.objects.filter(reviewed_by=user).update(reviewed_by=None)
    TeacherProgressCorrectionRequest.objects.filter(reviewed_by=user).update(reviewed_by=None)
    TeacherProgressCorrectionRequest.objects.filter(requested_by=user).delete()
    CorrectionRequest.objects.filter(requested_by=user).delete()
    # Delete teacher module data owned by this user.
    assignments = TeacherCourseAssignment.objects.filter(teacher=user)
    TeacherCourseProgress.objects.filter(assignment__in=assignments).delete()
    assignments.delete()
    # Delete copy-checking records tied to this user if admin chooses to delete the account.
    CopyCheckRecord.objects.filter(Q(entered_by=user) | Q(actual_checker_user=user)).delete()
    DailyBackup.objects.filter(triggered_by=user).update(triggered_by=None)
    profile = getattr(user, 'checker_profile', None)
    if profile:
        profile.delete()
    username = user.username
    user.delete()
    return username


def _delete_students(student_ids):
    students = Student.objects.filter(id__in=student_ids)
    CopyCheckRecord.objects.filter(student__in=students).delete()
    count = students.count()
    students.delete()
    return count


def _delete_classrooms(class_ids):
    classrooms = ClassRoom.objects.filter(id__in=class_ids)
    count = classrooms.count()
    class_subjects = ClassSubject.objects.filter(classroom__in=classrooms)
    assignments = TeacherCourseAssignment.objects.filter(class_subject__in=class_subjects)
    TeacherCourseProgress.objects.filter(assignment__in=assignments).delete()
    assignments.delete()
    CopyCheckRecord.objects.filter(Q(classroom__in=classrooms) | Q(student__classroom__in=classrooms) | Q(class_subject__in=class_subjects)).delete()
    Student.objects.filter(classroom__in=classrooms).delete()
    ClassSubjectChapter.objects.filter(class_subject__in=class_subjects).delete()
    class_subjects.delete()
    classrooms.delete()
    return count


def _delete_subjects(subject_ids):
    subjects = Subject.objects.filter(id__in=subject_ids)
    count = subjects.count()
    class_subjects = ClassSubject.objects.filter(subject__in=subjects)
    assignments = TeacherCourseAssignment.objects.filter(class_subject__in=class_subjects)
    TeacherCourseProgress.objects.filter(assignment__in=assignments).delete()
    assignments.delete()
    CopyCheckRecord.objects.filter(class_subject__in=class_subjects).delete()
    ClassSubjectChapter.objects.filter(class_subject__in=class_subjects).delete()
    class_subjects.delete()
    subjects.delete()
    return count


def _delete_weeks(progress_ids):
    rows = TeacherCourseProgress.objects.filter(id__in=progress_ids)
    count = rows.count()
    rows.delete()
    return count


def _delete_teachers(user_ids, current_user):
    teachers = User.objects.filter(id__in=user_ids, checker_profile__role=UserProfile.ROLE_TEACHER).exclude(id=current_user.id)
    count = teachers.count()
    for teacher in list(teachers):
        _force_delete_user(teacher)
    return count

@app_admin_required
def admin_users(request):
    users = User.objects.select_related('checker_profile').order_by('checker_profile__role', 'username')
    admin_checker_users = users.filter(Q(checker_profile__role=UserProfile.ROLE_ADMIN) | Q(checker_profile__role=UserProfile.ROLE_CHECKER) | Q(checker_profile__isnull=True))
    teacher_users = users.filter(checker_profile__role=UserProfile.ROLE_TEACHER)
    return render(request, 'checker/admin_users.html', {
        'admin_checker_users': admin_checker_users,
        'teacher_users': teacher_users,
    })


@app_admin_required
def admin_user_edit(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    profile, _created = UserProfile.objects.get_or_create(user=user, defaults={'display_name': user.username})
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            user.is_active = profile.is_active_checker
            user.save(update_fields=['is_active'])
            log_action(request.user, 'EDIT_USER_PROFILE', profile, f'Edited profile for {user.username}')
            messages.success(request, 'User details updated.')
            return redirect('admin_users')
    else:
        form = UserProfileForm(instance=profile)
    return render(request, 'checker/admin_user_edit.html', {'form': form, 'edited_user': user})


@app_admin_required
def admin_user_reset_password(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    profile, _created = UserProfile.objects.get_or_create(user=user, defaults={'display_name': user.username})
    if request.method == 'POST':
        form = PasswordResetByAdminForm(request.POST)
        if form.is_valid():
            password = form.cleaned_data['new_password']
            user.set_password(password)
            user.save(update_fields=['password'])
            if not user.check_password(password):
                messages.error(request, 'Password could not be verified after saving. Please try again.')
                return redirect('admin_user_reset_password', user_id=user.id)
            if form.cleaned_data['keep_visible_note']:
                profile.initial_password_note = password
                profile.save(update_fields=['initial_password_note'])
            if user.pk == request.user.pk:
                update_session_auth_hash(request, user)
            log_action(request.user, 'RESET_PASSWORD', profile, f'Reset password for {user.username}')
            ensure_daily_backup(reason='Admin reset a user password', user=request.user)
            messages.success(request, 'Login password reset successfully. The visible password note was updated too.')
            return redirect('admin_users')
    else:
        form = PasswordResetByAdminForm()
    return render(request, 'checker/admin_user_reset_password.html', {'form': form, 'edited_user': user})


@app_admin_required
def admin_user_delete(request, user_id):
    user = get_object_or_404(User.objects.select_related('checker_profile'), pk=user_id)
    profile = getattr(user, 'checker_profile', None)

    if user.pk == request.user.pk:
        messages.error(request, 'You cannot delete the account you are currently using.')
        return redirect('admin_users')

    if request.method == 'POST':
        form = UserDeleteConfirmForm(request.POST)
        if form.is_valid():
            display = profile.display_name if profile else user.username
            ensure_daily_backup(reason='Backup before deleting user', user=request.user)
            with transaction.atomic():
                username = _force_delete_user(user)
            log_action(request.user, 'DELETE_USER', None, f'Force deleted user {username} ({display})')
            messages.success(request, f'User {username} was deleted.')
            return redirect('admin_users')
    else:
        form = UserDeleteConfirmForm()

    return render(request, 'checker/admin_user_delete.html', {'form': form, 'deleted_user': user, 'profile': profile})


@app_admin_required
def backups_admin(request):
    backups = DailyBackup.objects.select_related('triggered_by')[:50]
    return render(request, 'checker/backups_admin.html', {
        'backups': backups,
        'upload_form': BackupUploadForm(),
    })


@app_admin_required
@require_POST
def backup_now(request):
    backup = ensure_daily_backup(reason='Manual admin backup', user=request.user)
    if not backup:
        messages.error(request, 'Backup could not be created.')
        return redirect('backups_admin')
    path = get_backup_path(backup)
    if not path.exists():
        messages.error(request, 'Backup record was created, but the file is not available on this server.')
        return redirect('backups_admin')
    return FileResponse(path.open('rb'), as_attachment=True, filename=backup.file_name)


@app_admin_required
@require_POST
def backup_upload(request):
    form = BackupUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, 'Please upload a valid JSON backup file.')
        return redirect('backups_admin')
    uploaded = form.cleaned_data['file']
    ensure_daily_backup(reason='Backup before uploaded restore', user=request.user)
    with tempfile.NamedTemporaryFile(suffix='.json', delete=False) as tmp:
        for chunk in uploaded.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name
    try:
        call_command('loaddata', tmp_path, verbosity=0)
    except Exception as exc:
        messages.error(request, f'Backup upload failed: {exc}')
        return redirect('backups_admin')
    log_action(request.user, 'UPLOAD_BACKUP', None, f'Uploaded backup file {uploaded.name}')
    ensure_daily_backup(reason='After uploaded backup restore', user=request.user)
    messages.success(request, 'Backup file uploaded and restored.')
    return redirect('backups_admin')


@app_admin_required
def backup_download(request, backup_id):
    backup = get_object_or_404(DailyBackup, pk=backup_id)
    path = get_backup_path(backup)
    if not path.exists():
        raise Http404('Backup file is not available on this server.')
    return FileResponse(path.open('rb'), as_attachment=True, filename=backup.file_name)



def _assignment_summary(assignment):
    total = assignment.progress_rows.count()
    completed = assignment.progress_rows.filter(status=TeacherCourseProgress.STATUS_COMPLETED).count()
    not_completed = max(total - completed, 0)
    return {
        'assignment': assignment,
        'total': total,
        'completed': completed,
        'not_completed': not_completed,
        'progress': _progress_percent(completed, total),
    }


def _teacher_assignment_queryset():
    return TeacherCourseAssignment.objects.filter(
        is_active=True,
        teacher__is_active=True,
        teacher__checker_profile__role=UserProfile.ROLE_TEACHER,
        teacher__checker_profile__is_active_checker=True,
        class_subject__is_active=True,
        class_subject__classroom__is_active=True,
        class_subject__subject__is_active=True,
    ).select_related(
        'teacher__checker_profile',
        'class_subject__classroom',
        'class_subject__subject',
    )


def _assignment_next_week_no(assignment):
    latest = assignment.progress_rows.aggregate(max_week=Max('week_no'))['max_week'] or 0
    return latest + 1


def _issue_week_rows_for_assignments(assignments):
    rows = []
    for assignment in assignments:
        next_week_no = _assignment_next_week_no(assignment)
        rows.append({
            'assignment': assignment,
            'next_week_no': next_week_no,
            'teacher_name': _display_user(assignment.teacher),
            'teacher_username': assignment.teacher.username,
            'classroom': assignment.classroom,
            'subject': assignment.subject,
            'existing_weeks': assignment.progress_rows.count(),
        })
    return rows


def _scope_filtered_assignments(form):
    assignments = _teacher_assignment_queryset()
    scope = form.cleaned_data['scope']
    if scope == IssueWeekForm.SCOPE_TEACHER:
        assignments = assignments.filter(teacher=form.cleaned_data['teacher'])
    elif scope == IssueWeekForm.SCOPE_CLASS:
        assignments = assignments.filter(class_subject__classroom=form.cleaned_data['classroom'])
    elif scope == IssueWeekForm.SCOPE_COURSE:
        assignments = assignments.filter(class_subject=form.cleaned_data['class_subject'])
    return assignments.order_by(
        'teacher__checker_profile__display_name',
        'teacher__username',
        'class_subject__classroom__name',
        'class_subject__classroom__section',
        'class_subject__subject__name',
    )


@app_admin_required
def admin_user_create(request):
    if request.method == 'POST':
        form = UserCreateByAdminForm(request.POST)
        if form.is_valid():
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
                is_active=form.cleaned_data['is_active'],
            )
            profile = UserProfile.objects.create(
                user=user,
                display_name=form.cleaned_data['display_name'],
                role=form.cleaned_data['role'],
                initial_password_note=form.cleaned_data['password'],
                is_active_checker=form.cleaned_data['is_active'],
            )
            log_action(request.user, 'CREATE_USER', profile, f'Created {profile.get_role_display()} user {user.username}')
            ensure_daily_backup(reason='Admin created a user', user=request.user)
            messages.success(request, 'User created successfully.')
            return redirect('admin_users')
    else:
        form = UserCreateByAdminForm()
    return render(request, 'checker/admin_user_create.html', {'form': form})


@teacher_required
def teacher_course_detail(request, assignment_id):
    assignment = get_object_or_404(
        TeacherCourseAssignment.objects.select_related('class_subject__classroom', 'class_subject__subject'),
        pk=assignment_id,
        teacher=request.user,
        is_active=True,
    )
    pending_rows = _teacher_pending_progress_rows(
        assignment.progress_rows.filter(status=TeacherCourseProgress.STATUS_NOT_COMPLETED)
    )
    completed_rows = assignment.progress_rows.filter(status=TeacherCourseProgress.STATUS_COMPLETED).order_by('-week_no')
    return render(request, 'checker/teacher_course_detail.html', {
        'assignment': assignment,
        'pending_rows': pending_rows,
        'completed_rows': completed_rows,
        'summary': _assignment_summary(assignment),
    })


@teacher_required
def teacher_progress_edit(request, progress_id):
    progress = get_object_or_404(TeacherCourseProgress, pk=progress_id, assignment__teacher=request.user)
    messages.error(request, 'Teachers cannot edit week rows directly. Enter detail when marking a pending week completed.')
    return redirect('teacher_course_detail', assignment_id=progress.assignment_id)


@teacher_required
@require_POST
def teacher_progress_delete(request, progress_id):
    progress = get_object_or_404(TeacherCourseProgress, pk=progress_id, assignment__teacher=request.user)
    messages.error(request, 'Teachers cannot delete week rows. Ask admin if a row was issued by mistake.')
    return redirect('teacher_course_detail', assignment_id=progress.assignment_id)


@teacher_required
@require_POST
def teacher_progress_complete(request, progress_id):
    progress = get_object_or_404(TeacherCourseProgress, pk=progress_id, assignment__teacher=request.user)
    if not progress.can_edit:
        messages.info(request, 'This row is already completed/locked.')
        next_url = request.POST.get('next')
        if next_url:
            return redirect(next_url)
        return redirect('teacher_course_detail', assignment_id=progress.assignment_id)
    form = TeacherCompleteForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'Please enter detail before marking this week as completed.')
        return redirect(request.POST.get('next') or reverse('teacher_course_detail', kwargs={'assignment_id': progress.assignment_id}))
    progress.detail = form.cleaned_data['detail'].strip()
    progress.status = TeacherCourseProgress.STATUS_COMPLETED
    progress.save(update_fields=['detail', 'status', 'locked', 'completed_at', 'updated_at'])
    log_action(request.user, 'COMPLETE_TEACHER_PROGRESS', progress, f'Completed week {progress.week_no}')
    ensure_daily_backup(reason='Teacher completed a progress row', user=request.user)
    messages.success(request, f'Week {progress.week_no} marked completed and locked.')
    next_url = request.POST.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('teacher_course_detail', assignment_id=progress.assignment_id)


@teacher_required
def teacher_quick_update(request):
    pending_rows = TeacherCourseProgress.objects.filter(
        assignment__teacher=request.user,
        status=TeacherCourseProgress.STATUS_NOT_COMPLETED,
        assignment__is_active=True,
    ).select_related('assignment__class_subject__classroom', 'assignment__class_subject__subject')
    return render(request, 'checker/teacher_quick_update.html', {
        'pending_rows': _teacher_pending_progress_rows(pending_rows),
    })



@app_admin_required
def issue_week_admin(request):
    preview_rows = None
    created_rows = None
    skipped_rows = None

    if request.method == 'POST':
        form = IssueWeekForm(request.POST)
        if form.is_valid():
            assignments = _scope_filtered_assignments(form)
            issue_rows = _issue_week_rows_for_assignments(assignments)

            if not issue_rows:
                messages.error(
                    request,
                    'No active teacher-course assignment matched this scope. Go to School Setup → Teacher Assignments and make sure the teacher has an active class-subject assignment.'
                )
                preview_rows = []
            elif 'preview' in request.POST:
                preview_rows = issue_rows
                messages.info(request, f'Preview ready. {len(preview_rows)} active assignment(s) matched this scope.')
            elif request.POST.get('confirm') == 'yes':
                created_rows = []
                skipped_rows = []
                with transaction.atomic():
                    locked_assignments = _scope_filtered_assignments(form).select_for_update()
                    for assignment in locked_assignments:
                        next_week_no = _assignment_next_week_no(assignment)
                        obj, created = TeacherCourseProgress.objects.get_or_create(
                            assignment=assignment,
                            week_no=next_week_no,
                            defaults={
                                'detail': '',
                                'status': TeacherCourseProgress.STATUS_NOT_COMPLETED,
                                'locked': False,
                            },
                        )
                        row = {
                            'assignment': assignment,
                            'week_no': next_week_no,
                            'teacher_name': _display_user(assignment.teacher),
                            'teacher_username': assignment.teacher.username,
                            'classroom': assignment.classroom,
                            'subject': assignment.subject,
                        }
                        if created:
                            created_rows.append(row)
                            log_action(request.user, 'ISSUE_TEACHER_WEEK', obj, f'Issued week {next_week_no} for {assignment}')
                        else:
                            skipped_rows.append(row)

                if created_rows:
                    ensure_daily_backup(reason='Admin issued new teacher week rows', user=request.user)
                    messages.success(request, f'Issued new week rows. Created: {len(created_rows)}. Skipped existing: {len(skipped_rows)}.')
                else:
                    messages.warning(request, 'No week rows were created. Check whether active teacher-course assignments exist for this scope.')
                preview_rows = _issue_week_rows_for_assignments(_scope_filtered_assignments(form))
    else:
        form = IssueWeekForm(initial={'scope': IssueWeekForm.SCOPE_ALL})

    total_active_assignments = _teacher_assignment_queryset().count()

    return render(request, 'checker/issue_week_admin.html', {
        'form': form,
        'preview_rows': preview_rows,
        'created_rows': created_rows,
        'skipped_rows': skipped_rows,
        'total_active_assignments': total_active_assignments,
    })


@app_admin_required
def teacher_assignments_admin(request):
    if request.method == 'POST':
        form = TeacherCourseAssignmentForm(request.POST)
        if form.is_valid():
            assignment = form.save()
            log_action(request.user, 'SAVE_TEACHER_ASSIGNMENT', assignment, f'Saved teacher assignment {assignment}')
            ensure_daily_backup(reason='Admin saved teacher assignment', user=request.user)
            messages.success(request, 'Teacher assignment saved.')
            return redirect('teacher_assignments_admin')
    else:
        form = TeacherCourseAssignmentForm()

    assignments = TeacherCourseAssignment.objects.select_related(
        'teacher__checker_profile', 'class_subject__classroom', 'class_subject__subject'
    ).order_by('class_subject__classroom__name', 'class_subject__classroom__section', 'class_subject__subject__name')
    return render(request, 'checker/teacher_assignments_admin.html', {'form': form, 'assignments': assignments})


@app_admin_required
@require_POST
def teacher_assignment_deactivate(request, assignment_id):
    assignment = get_object_or_404(TeacherCourseAssignment, pk=assignment_id)
    assignment.is_active = False
    assignment.save(update_fields=['is_active', 'updated_at'])
    log_action(request.user, 'DEACTIVATE_TEACHER_ASSIGNMENT', assignment, f'Deactivated {assignment}')
    ensure_daily_backup(reason='Admin deactivated teacher assignment', user=request.user)
    messages.success(request, 'Teacher assignment deactivated.')
    return redirect('teacher_assignments_admin')


@app_admin_required
def teacher_progress_classes_admin(request):
    classes = ClassRoom.objects.filter(is_active=True).annotate(
        assignment_count=Count('class_subjects__teacher_assignments', filter=Q(class_subjects__teacher_assignments__is_active=True)),
    ).order_by('name', 'section')
    return render(request, 'checker/teacher_progress_classes_admin.html', {'classes': classes})


@app_admin_required
def teacher_progress_class_detail_admin(request, class_id):
    classroom = get_object_or_404(ClassRoom, pk=class_id)
    assignments = TeacherCourseAssignment.objects.filter(
        class_subject__classroom=classroom,
        is_active=True,
    ).select_related('teacher__checker_profile', 'class_subject__classroom', 'class_subject__subject')
    rows = [_assignment_summary(assignment) for assignment in assignments]
    return render(request, 'checker/teacher_progress_class_detail_admin.html', {'classroom': classroom, 'rows': rows})


@app_admin_required
def teacher_progress_teachers_admin(request):
    teachers = User.objects.filter(
        checker_profile__role=UserProfile.ROLE_TEACHER,
        checker_profile__is_active_checker=True,
    ).select_related('checker_profile').annotate(
        assignment_count=Count('teacher_course_assignments', filter=Q(teacher_course_assignments__is_active=True)),
    ).order_by('checker_profile__display_name', 'username')
    return render(request, 'checker/teacher_progress_teachers_admin.html', {'teachers': teachers})


@app_admin_required
def teacher_progress_teacher_detail_admin(request, user_id):
    teacher = get_object_or_404(User.objects.select_related('checker_profile'), pk=user_id, checker_profile__role=UserProfile.ROLE_TEACHER)
    assignments = TeacherCourseAssignment.objects.filter(
        teacher=teacher,
        is_active=True,
    ).select_related('teacher__checker_profile', 'class_subject__classroom', 'class_subject__subject')
    rows = [_assignment_summary(assignment) for assignment in assignments]
    return render(request, 'checker/teacher_progress_teacher_detail_admin.html', {'teacher': teacher, 'rows': rows})


@app_admin_required
def teacher_progress_assignment_detail_admin(request, assignment_id):
    assignment = get_object_or_404(
        TeacherCourseAssignment.objects.select_related('teacher__checker_profile', 'class_subject__classroom', 'class_subject__subject'),
        pk=assignment_id,
    )
    rows = assignment.progress_rows.all()
    return render(request, 'checker/teacher_progress_assignment_detail_admin.html', {
        'assignment': assignment,
        'rows': rows,
        'summary': _assignment_summary(assignment),
    })


@app_admin_required
def teacher_progress_row_edit_admin(request, progress_id):
    progress = get_object_or_404(
        TeacherCourseProgress.objects.select_related(
            'assignment__teacher__checker_profile',
            'assignment__class_subject__classroom',
            'assignment__class_subject__subject',
        ),
        pk=progress_id,
    )
    if request.method == 'POST':
        form = AdminTeacherProgressEditForm(request.POST)
        if form.is_valid():
            old_detail = progress.detail
            old_status = progress.status
            progress.detail = form.cleaned_data['detail']
            progress.status = form.cleaned_data['status']
            progress.save(update_fields=['detail', 'status', 'locked', 'completed_at', 'updated_at'])
            note = form.cleaned_data.get('admin_note') or ''
            log_action(
                request.user,
                'ADMIN_EDIT_TEACHER_PROGRESS',
                progress,
                f'Admin edited week {progress.week_no}. Old status: {old_status}. New status: {progress.status}. Note: {note}. Old detail: {old_detail[:120]}',
            )
            ensure_daily_backup(reason='Admin edited teacher progress row', user=request.user)
            messages.success(request, 'Teacher progress row updated by admin.')
            return redirect('teacher_progress_assignment_detail_admin', assignment_id=progress.assignment_id)
    else:
        form = AdminTeacherProgressEditForm(initial={
            'detail': progress.detail,
            'status': progress.status,
        })
    return render(request, 'checker/teacher_progress_row_edit_admin.html', {
        'form': form,
        'progress': progress,
        'assignment': progress.assignment,
    })




@app_admin_required
def teacher_progress_export(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Teacher Progress'
    ws.append(['Teacher', 'Class', 'Subject', 'Week No', 'Detail', 'Status', 'Created', 'Completed'])
    rows = TeacherCourseProgress.objects.select_related(
        'assignment__teacher__checker_profile',
        'assignment__class_subject__classroom',
        'assignment__class_subject__subject',
    ).order_by(
        'assignment__teacher__checker_profile__display_name',
        'assignment__class_subject__classroom__name',
        'assignment__class_subject__classroom__section',
        'assignment__class_subject__subject__name',
        'week_no'
    )
    for row in rows:
        ws.append([
            _display_user(row.assignment.teacher),
            str(row.assignment.classroom),
            row.assignment.subject.name,
            row.week_no,
            row.detail,
            row.get_status_display(),
            timezone.localtime(row.created_at).strftime('%Y-%m-%d %H:%M') if row.created_at else '',
            timezone.localtime(row.completed_at).strftime('%Y-%m-%d %H:%M') if row.completed_at else '',
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="teacher_progress.xlsx"'
    wb.save(response)
    return response


def _delete_data_items(category, ids, current_user):
    if category == 'teachers':
        return _delete_teachers(ids, current_user)
    if category == 'weeks':
        return _delete_weeks(ids)
    if category == 'subjects':
        return _delete_subjects(ids)
    if category == 'classes':
        return _delete_classrooms(ids)
    if category == 'students':
        return _delete_students(ids)
    if category == 'records':
        qs = CopyCheckRecord.objects.filter(id__in=ids)
        count = qs.count()
        qs.delete()
        return count
    return 0


def _delete_data_queryset(category):
    if category == 'teachers':
        return User.objects.filter(checker_profile__role=UserProfile.ROLE_TEACHER).select_related('checker_profile').order_by('checker_profile__display_name', 'username')
    if category == 'weeks':
        return TeacherCourseProgress.objects.select_related(
            'assignment__teacher__checker_profile',
            'assignment__class_subject__classroom',
            'assignment__class_subject__subject',
        ).order_by('-week_no', 'assignment__class_subject__classroom__name', 'assignment__class_subject__subject__name')
    if category == 'subjects':
        return Subject.objects.all().order_by('name')
    if category == 'classes':
        return ClassRoom.objects.all().order_by('name', 'section')
    if category == 'students':
        return Student.objects.select_related('classroom').order_by('classroom__name', 'classroom__section', 'roll_no', 'full_name')
    if category == 'records':
        return _checking_record_queryset().order_by('-locked_at', '-id')[:1000]
    return []


def _delete_data_label(category, obj):
    if category == 'teachers':
        return f'{_display_user(obj)} ({obj.username})'
    if category == 'weeks':
        return f'Week {obj.week_no} — {obj.assignment.classroom} — {obj.assignment.subject.name} — {_display_user(obj.assignment.teacher)}'
    if category == 'subjects':
        return obj.name
    if category == 'classes':
        return str(obj)
    if category == 'students':
        return f'{obj.classroom} — Roll {obj.roll_no} — {obj.full_name}'
    if category == 'records':
        return f'{obj.locked_at:%Y-%m-%d} — {obj.classroom} — {obj.student.full_name} — {obj.subject.name} — {obj.get_status_display()}'
    return str(obj)


@app_admin_required
def delete_data_admin(request):
    category_choices = [
        ('teachers', 'Teachers'),
        ('weeks', 'Teacher week rows'),
        ('subjects', 'Subjects'),
        ('classes', 'Classes'),
        ('students', 'Students'),
        ('records', 'Checking records'),
    ]
    category = request.GET.get('category') or request.POST.get('category') or 'teachers'
    valid_categories = {key for key, _label in category_choices}
    if category not in valid_categories:
        category = 'teachers'

    items = list(_delete_data_queryset(category))
    item_rows = [{'id': item.id, 'label': _delete_data_label(category, item), 'object': item} for item in items]

    if request.method == 'POST':
        mode = request.POST.get('mode') or 'selected'
        selected_ids = [int(item_id) for item_id in request.POST.getlist('selected_ids') if item_id.isdigit()]
        if mode == 'all':
            selected_ids = [row['id'] for row in item_rows]
        if category == 'teachers':
            selected_ids = [item_id for item_id in selected_ids if item_id != request.user.id]
        selected_labels = [row['label'] for row in item_rows if row['id'] in selected_ids]
        if not selected_ids:
            messages.error(request, 'Select at least one item to delete.')
            return redirect(f'{reverse("delete_data_admin")}?category={category}')
        if request.POST.get('confirm_text') != 'DELETE':
            return render(request, 'checker/delete_data_confirm.html', {
                'category': category,
                'category_choices': category_choices,
                'selected_ids': selected_ids,
                'selected_labels': selected_labels,
                'mode': mode,
            })
        ensure_daily_backup(reason=f'Backup before deleting {category}', user=request.user)
        try:
            with transaction.atomic():
                count = _delete_data_items(category, selected_ids, request.user)
        except (ProtectedError, IntegrityError) as exc:
            messages.error(request, f'Deletion was blocked because related records could not be removed safely: {exc}')
            return redirect(f'{reverse("delete_data_admin")}?category={category}')
        log_action(request.user, 'DELETE_DATA', None, f'Deleted {count} item(s) from {category}')
        messages.success(request, f'Deleted {count} item(s) from {dict(category_choices).get(category, category)}.')
        return redirect(f'{reverse("delete_data_admin")}?category={category}')

    return render(request, 'checker/delete_data_admin.html', {
        'category': category,
        'category_choices': category_choices,
        'item_rows': item_rows,
    })
